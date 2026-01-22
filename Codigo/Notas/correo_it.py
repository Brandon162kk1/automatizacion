#-- Imports --
import base64
import json
import requests
import logging
import os
import signal
import time
import sys
import pandas as pd
#-- Froms --
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from Birlik.cancelar_cuotas import url_cuotas
from GoogleChrome.chromeDriver import get_fecha_actual,get_dia,get_mes

#------------CORREO IT DESTINATARIOS---------------
remitente = "it@birlik.com.pe"
asunto = "Automatización del Proceso de Validación de Cuotas Canceladas y Registro en Birlik 365"
mensaje = """
<p>Soy <b>CobraBirlik</b>, tu asistente automatizado de cobranza.</p>

<p>Acabo de ejecutar el proceso de cancelación de cuotas para las compañías <b>Sanitas Crecer</b>, <b>Sanitas Protecta</b> y <b>La Positiva</b> en los siguientes ramos:</p>

<ul>
  <li><b>Salud</b></li>
  <li><b>Pensión</b></li>
  <li><b>Vida Ley</b></li>
</ul> 

<p>Por favor, revisar el archivo Excel con el detalle de resultados y cualquier observación indicada me hacen llegar por este medio para continuar con la gestión o validación correspondiente.</p>

<p>Este proceso fue ejecutado automáticamente como parte de la automatización operativa diaria de Birlik.</p>

<p><b>Consideraciones:</b></p>

<p>Las últimas 5 o 6 filas que se agregaron al reporte del Excel nos darán una información valiosa para saber cómo se está manejando el flujo de cancelación de cuotas.</p>

<p>Se explicará de manera detallada cada columna agregada:</p>

<ul>
  <li><b>Columna 'Importe':</b> Se validará que los importes de la compañía y nuestro Sistema Birlik 365 coincidan.</li>
  <li><b>Columna 'Sunat' :</b> Es el resultado en la página de SUNAT si estos documentos son válidos.</li>
  <li><b>Columna 'Birlik' :</b> Se registrará en la plataforma BIRLIK 365 si la cuota tiene estado "Cancelado" o "Abonada".</li>
  <li><b>Columna 'OCR' :</b> Se indica si se utilizó el servicio OCR para leer las fechas de emisión.</li>
  <li><b>Columna 'Estado' :</b> Estado de la compañía con respecto a nuestra Plataforma.</li>
  <li><b>Columna 'Acción' :</b> Ingresar a la acción señelada a través de un hipervínculo para tomar las acciones necesarias.</li>
</ul> 

<p><b>Atentamente,</b><br>
RobotBirlik – Automatización Inteligente</p>
"""
#----- PRUEBAS DE CORREOS-----------
destinatarios_to =["brandon.rodriguez@jishu.com.pe"]
destinatarios_cc =["bruno.zevallos@jishu.com.pe"]
#------Carpetas de Descargas y Volumen del Docker----------
carpeta_descargas = "Downloads"
ruta_carpeta_descargas = f"/app/{carpeta_descargas}"
#------ Reporte de Cuotas Diarias -------
nom_carp_principal= f"Reporte_Cuotas_Diarias_{get_fecha_actual()}"
# Se crea carpetas dentro de Downloads,Ejemplo --> :/app/Downloads/Reporte_Cuotas_Diarias_2025-07-21
carpeta_principal = os.path.join(ruta_carpeta_descargas, nom_carp_principal)
salida_reporte_final= 'Reporte_Final_Cuotas.xlsx'
#Ejemplo --> :/app/Downloads/Reporte_Cuotas_Diarias_2025-07-21/Reporte_Final_Cuotas.xlsx
ruta_maestro = os.path.join(carpeta_principal, salida_reporte_final)

#---------Configuración de la cuenta en Azure-----
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
AUTHORITY = f'https://login.microsoftonline.com/{TENANT_ID}'
SCOPE = [os.getenv("SCOPE")]
EMAIL_ACCOUNT = os.getenv("remitente")
#-----MS Graph API URL para obtener correos-------
GRAPH_API_URL = 'https://graph.microsoft.com/v1.0/users/{}/messages'.format(EMAIL_ACCOUNT)

def enviarCorreoxEjecutivo():

    # Carpeta para guardar los reportes
    nombre_carpeta_anular = "Cuotas_x_Anular_de_Ejecutivos"
    subcarpeta_CuotasAnular = os.path.join(carpeta_principal, nombre_carpeta_anular)
    os.makedirs(subcarpeta_CuotasAnular, exist_ok=True)

    # Leer todas las hojas
    xls = pd.ExcelFile(ruta_maestro)
    cuotas_por_ejecutivo = {}

    for nombre_hoja in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=nombre_hoja)

        # Filtrar cuotas por anular (ajusta el criterio según tu lógica)
        df_anular = df[df["Estado"].str.upper().str.contains("ANUL",na=False)]

        if df_anular.empty:
            continue

        # Agrupar por ejecutivo
        for ejecutivo, grupo_df in df_anular.groupby("ejecutivoResponsable"):
            if ejecutivo not in cuotas_por_ejecutivo:
                cuotas_por_ejecutivo[ejecutivo] = grupo_df
            else:
                cuotas_por_ejecutivo[ejecutivo] = pd.concat([cuotas_por_ejecutivo[ejecutivo], grupo_df])

    # Procesar y enviar correos
    for ejecutivo, df_final in cuotas_por_ejecutivo.items():
        if df_final.empty:
            continue

        nombre = ejecutivo.split("@")[0].split(".")[0].capitalize()
        nombre_archivo = f"{nombre}_Cuotas_Anular.xlsx".replace(" ", "_")
        ruta_archivo = os.path.join(subcarpeta_CuotasAnular, nombre_archivo)

        df_final.to_excel(ruta_archivo, index=False)

        # --- Mapeo de fk_Compania ---
        mapeo_companias = {
            31: "Protecta",
            12: "Positiva",
            13: "Positiva",
            14: "Positiva",
            36: "Positiva",
            38: "Positiva",
            5:  "Crecer",
            29: "Sanitas"
        }

        # Agregar hipervínculos a la columna "Acción"
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        # --- Buscar columna fk_Compania ---
        columna_compania_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "fK_Compania":
                columna_compania_idx = idx
                break

        if columna_compania_idx:
            for fila_idx in range(2, ws.max_row + 1):
                celda = ws.cell(row=fila_idx, column=columna_compania_idx)
                try:
                    valor = int(celda.value)
                    if valor in mapeo_companias:
                        celda.value = mapeo_companias[valor]
                except Exception as e:
                    logging.warning(f"⚠️ Error al reemplazar fk_Compania en fila {fila_idx}: {e}")
        else:
            logging.warning("⚠️ No se encontró la columna 'fk_Compania'")

        
        # Busca la columna "Acción"
        columna_accion_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Acción":
                columna_accion_idx = idx
                break

        if not columna_accion_idx:
            logging.info("❌ No se encontró la columna 'Acción'")
        else:
            # Recorres cada fila del archivo guardado
            for fila_idx, (_, row) in enumerate(df_final.iterrows(), start=2):
            #for fila_idx, (_, row) in enumerate(grupo_df.iterrows(), start=2):
                fk_cliente = row['fk_Cliente']
                celda_accion = ws.cell(row=fila_idx, column=columna_accion_idx)

                url = f"{url_cuotas}{fk_cliente}"
                # Pones un hipervínculo real, no texto plano
                celda_accion.value = f'=HYPERLINK("{url}", "Anular Cuota")'
                celda_accion.font = Font(color="FF0000")

        # Ajustar ancho de columnas
        for col in ws.columns:
             max_length = 0
             col_letter = col[0].column_letter
             for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                adjusted_width = max_length + 2
                ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(ruta_archivo)

        # Enviar correo
        destinatarios1 = [ejecutivo]#destinatarios_cc
        destinatarios2 = ["brandon.rodriguez@jishu.com.pe"]
        asunto_ejecutivo= f"Reporte de Cuotas por Anular - {datetime.now().strftime('%d/%m/%Y')}"
        if len(df_final) == 1:
            auto_msj = "cuota pendiente"
        else:
            auto_msj = "cuotas pendientes"
        mensaje_ejecutivo = f"Estimad@ {nombre},<br><br>Adjuntamos el reporte detallado de {len(df_final)} {auto_msj} por anular.<br>"
        mensaje_ejecutivo += "<br>Saludos.<br>Área de Cobranza"
        
        time.sleep(2)  # Esperar un poco antes de enviar el correo  

        print(f"-- {ejecutivo} tiene {len(df_final)} {auto_msj} por anular.")
        enviarCorreoIT(destinatarios1, destinatarios2, asunto_ejecutivo, mensaje_ejecutivo, lista_adjuntos=[ruta_archivo])

def guardar_json_a_excel(json_data, ruta_salida):
    try:
        df = pd.DataFrame(json_data)
        df.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"✅ Excel creado en {ruta_salida}")
    except Exception as e:
        print(f"❌ Error guardando Excel: {e}")

def formato_correos(lista):
    return [{"emailAddress": {"address": correo}} for correo in lista]

def enviarCorreoIT(destinatarios,destinatarios_copia,asunto, mensaje_html, lista_adjuntos):

    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")
    tenant_id = os.getenv("tenant_id")
    token_endpoint = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
 
    # Paso 1: Obtener el token
    token_data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default"
    }
    token_response = requests.post(token_endpoint, data=token_data)
    access_token = token_response.json().get("access_token")
 
    if not access_token:
        print("❌ No se pudo obtener el token.")
        return
 
    # Paso 2: Construir el cuerpo del mensaje
    email_body = {
        "message": {
            "subject": asunto,
            "body": {
                "contentType": "HTML",
                "content": f"""
        <html>
        <body>
        {mensaje_html}
        </body>
        </html>
                """
            },
            "toRecipients": formato_correos(destinatarios),
            "ccRecipients": formato_correos(destinatarios_copia),
            #"attachments": []
        },
        "saveToSentItems": "true"
    }
 
    attachments = []

    for archivo_path in lista_adjuntos:
        if os.path.exists(archivo_path):
            with open(archivo_path, "rb") as f:
                contenido_base64 = base64.b64encode(f.read()).decode('utf-8')
                attachments.append({
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": os.path.basename(archivo_path),
                    "contentBytes": contenido_base64,
                    "contentType": "application/octet-stream"
                })
        else:
            print(f"\n⚠️ Archivo no encontrado: {archivo_path}")
 
    # Solo agregar attachments si existen
    if attachments:
         email_body["message"]["attachments"] = attachments

    # Paso 3: Enviar el correo con Microsoft Graph
    url_envio = f"https://graph.microsoft.com/v1.0/users/{remitente}/sendMail"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }
    try:
        response = requests.post(
            url_envio,
            headers=headers,
            data=json.dumps(email_body),
            timeout=30
        )
        if response.status_code == 202:
            print(f"\n📧 Correo enviado correctamente a {destinatarios} y con copia a {destinatarios_copia}.")
        else:
            print("\n❌ Error al enviar correo:", response.status_code, response.text)
    except requests.exceptions.Timeout:
        print("\n⏰ El envío del correo tardó demasiado y fue cancelado (timeout).")
    except Exception as e:
        print(f"\n❌ Error inesperado al enviar el correo: {e}")

# def cambiarmetodo():

#     # --- Armando el nombre del log para ver los resultados del enviar Correo ---
#     nombre_salida = f"logs_Correo_{get_dia()}_{get_mes()}.txt"

#     carpeta_logs = os.path.join(carpeta_principal)  # carpeta_principal debe apuntar a Downloads/Reporte_Cuotas...
#     os.makedirs(carpeta_logs, exist_ok=True)  # ✅ Crea la carpeta si no existe

#     log_path = os.path.join(carpeta_principal, nombre_salida)

#     # --- Validar archivo maestro antes de abrir el log ---
#     if not os.path.exists(ruta_maestro):
#         print(f"❌ El archivo maestro no existe en: {ruta_maestro}, Finaliza todo la Automatización")
#         os.kill(1, signal.SIGTERM) # 🔹 Intenta apagar contenedor
#         sys.exit(1)                # 🔹 Cortar ejecución inmediatamente

#     # --- Redireccionar todo el stdout al archivo de log ---
#     original_stdout = sys.stdout  # Guarda referencia original a la consola
#     with open(log_path, "w", encoding="utf-8") as log_file:
#         #sys.stdout = log_file
#         sys.stdout = Tee(sys.stdout, log_file)

#         #Primero enviamos facturas a los clientes que no son de Factura Adelantada y que tenga el importe igual
#         # try:
#         #     print("⌛ Analizando Cuotas Canceladas para enviar Facturas al cliente")
#         #     analizarFacturasparaEnviarCliente(ruta_maestro)
#         # except Exception as e:
#         #     print(f"❌ Error al enviar Facturas al cliente, Detalle del error: {e}")
        
#         #print("----------------------")

#         #Segundo correo para cada ejecutivo que tiene cuotas pendientes por anular
#         try:
#             print("⌛ Enviando correo a cada ejecutivo de manera personalizada")
#             enviarCorreoxEjecutivo()
#         except Exception as e:
#             print(f"❌ Error al enviar correo a cada ejecutivo: {e}")

#         print("----------------------")
        
#         # Tercero, enivamos correo general al equipo gerencial
#         try:
#             print("⌛ Enviando el Reporte Final de las Cuotas para el equipo Gerencial")
#             enviarCorreoIT(destinatarios_to,destinatarios_cc,asunto, mensaje, lista_adjuntos=[ruta_maestro])
#         except Exception as e:
#             print(f"❌ Error al enviar correo gerencial: {e}")
#         finally:
#             sys.stdout = original_stdout  # Restaura la salida estándar (la consola)

#     # 🔻 Este es el lugar correcto para terminar el contenedor
#     os.kill(1, signal.SIGTERM)

# Metodo para guardar un Excel del correo IT que se reviso
def guardar_excel(ruta_carpeta, nombre_archivo, contenido_b64):
    """Guarda un archivo Excel en ruta_carpeta, renombrando si ya existe"""
    #os.makedirs(ruta_carpeta, exist_ok=True)
    nombre, extension = os.path.splitext(nombre_archivo)
    ruta_archivo = os.path.join(ruta_carpeta, nombre_archivo)
    #ruta_archivo = os.path.join(base_dir, "Downloads", nombre_archivo)
    contador = 1
    while os.path.exists(ruta_archivo):
        nuevo_nombre = f"{nombre} ({contador}){extension}"
        ruta_archivo = os.path.join(ruta_carpeta, nuevo_nombre)
        contador += 1

    with open(ruta_archivo, "wb") as f:
        f.write(base64.b64decode(contenido_b64))
    
    print(f"✅ Archivo Excel guardado: {ruta_archivo}")
    return ruta_archivo
import unicodedata
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def _sin_acentos(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def _texto_visible(celda) -> str:
    """Devuelve el texto que ve el usuario: si es fórmula HYPERLINK/HIPERVINCULO, extrae el segundo argumento."""
    v = celda.value
    if v is None:
        return ""
    if isinstance(v, str) and v.startswith("="):
        f = v.strip()
        f_upper = f.upper()
        if f_upper.startswith("=HYPERLINK(") or f_upper.startswith("=HIPERVINCULO("):
            # Captura todo lo dentro de paréntesis
            m = re.match(r"^=[A-ZÁÉÍÓÚÑ]+?\((.*)\)$", f, flags=re.IGNORECASE)
            if m:
                args = m.group(1)
                # Toma los textos entre comillas; el segundo suele ser el texto visible
                quoted = re.findall(r'"([^"]*)"', args)
                if quoted:
                    return quoted[1] if len(quoted) >= 2 else quoted[-1]
    return str(v)

def guardar_excel_con_formato(ruta_guardado,nombre_hoja):
    
    wb = load_workbook(ruta_guardado,data_only=False)  # 🔹 data_only=False para ver fórmulas)
    
    # Seleccionas la hoja correcta
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
    else:
        print(f"❌ No se encontró la hoja: {nombre_hoja}")
        return

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Pintar y subrayar si hay hipervínculo
    columna_accion = None
    columna_importe = None
    for cell in ws[1]:
        # if cell.value == "Acción":
        #     columna_accion = cell.column_letter
        #     break
        if cell.value == "Acción":
            columna_accion = cell.column_letter
        elif cell.value == "Importe":
            columna_importe = cell.column_letter

    # Recorrer filas
    for fila in range(2, ws.max_row + 1):
        # ---- Columna Acción ----
        if columna_accion:
            celda_accion = ws[f"{columna_accion}{fila}"]
            texto_accion = _texto_visible(celda_accion).strip()
            texto_norm_accion = _sin_acentos(texto_accion).lower()

            if texto_accion:  # Si no está vacía
                if texto_norm_accion == "sin observacion" or texto_norm_accion == "pagina web en mantenimiento" or texto_norm_accion == "factura enviada anteriormente":
                    celda_accion.font = Font(color="000000", underline=None) # Negro
                else:
                    celda_accion.font = Font(color="FF0000", underline="single") #Rojo Subrayado

        # ---- Columna Importe ----
        if columna_importe:
            celda_importe = ws[f"{columna_importe}{fila}"]
            texto_importe = str(celda_importe.value or "").strip().lower()

            if texto_importe in ("no coinciden", "no existe en la cia"):#== "no coinciden":
                celda_importe.font = Font(color="FF0000") #Rojo negrita(bold=True)

    wb.save(ruta_guardado)
    print(f"\n💾 Archivo guardado y formateado correctamente en:\n{ruta_guardado} -> Hoja: {nombre_hoja}")

def guardar_excel_con_formato_solo_ajustar_columnas(ruta_guardado,nombre_hoja):
    
    wb = load_workbook(ruta_guardado,data_only=False)  # 🔹 data_only=False para ver fórmulas)
    
    # Seleccionas la hoja correcta
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
    else:
        print(f"❌ No se encontró la hoja: {nombre_hoja}")
        return

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(ruta_guardado)
    print(f"\n✅ Archivo guardado y formateado correctamente en:\n{ruta_guardado} -> Hoja: {nombre_hoja}")

def guardar_json_a_excel(json_data, ruta_salida):
    try:
        df = pd.DataFrame(json_data)
        df.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"✅ Excel creado en {ruta_salida}")
    except Exception as e:
        print(f"❌ Error guardando Excel: {e}")